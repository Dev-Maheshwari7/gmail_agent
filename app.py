import imaplib
import email
import os
import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from langgraph.graph import StateGraph, END
from typing import TypedDict

# ==================== STATE DEFINITION ====================
class AgentState(TypedDict, total=False):
    """State shared across all agents"""
    gmail_account: str
    gmail_password: str
    email_ids: list
    email_data: list
    output_path: str
    message: str
    error: str

# ==================== HELPER FUNCTION ====================
def extract_email_address(email_string: str) -> str:
    """Extract email address from 'Name <email@address.com>' format"""
    if not email_string:
        return "N/A"
    
    # Try to extract email from <email@address.com> format
    match = re.search(r'<([^>]+)>', email_string)
    if match:
        return match.group(1)
    
    # If no angle brackets, check if it's already a valid email
    if '@' in email_string:
        return email_string.strip()
    
    return "N/A"

# ==================== AGENT 1: GMAIL EXTRACTOR ====================
def agent_gmail_extractor(state: AgentState) -> dict:
    """Extract sent emails from Gmail (limit to 10)"""
    print(f"\n[STEP 1] Connecting to Gmail...")
    sys.stdout.flush()
    
    try:
        gmail_account = state.get("gmail_account")
        gmail_password = state.get("gmail_password")
        
        if not gmail_account or not gmail_password:
            print(f"✗ Missing credentials")
            sys.stdout.flush()
            return {
                "error": "Missing credentials",
                "message": "Gmail account or password not provided"
            }
        
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(gmail_account, gmail_password)
        print(f"✓ Connected successfully")
        sys.stdout.flush()
        
        print(f"\n[STEP 2] Selecting Sent Mail folder...")
        imap.select('"[Gmail]/Sent Mail"')
        print(f"✓ Folder selected")
        sys.stdout.flush()
        
        print(f"\n[STEP 3] Fetching all sent emails...")
        status, messages = imap.search(None, "ALL")
        all_email_ids = messages[0].split() if messages[0] else []
        
        # LIMIT TO 10 EMAILS TO AVOID RECURSION
        email_ids = all_email_ids[-10:] if len(all_email_ids) > 10 else all_email_ids
        print(f"✓ Found {len(all_email_ids)} total, using {len(email_ids)} emails")
        sys.stdout.flush()
        
        imap.close()
        
        return {
            "email_ids": email_ids,
            "message": f"Found {len(email_ids)} emails"
        }
        
    except Exception as e:
        error_msg = str(e)
        print(f"✗ Connection failed: {error_msg}")
        sys.stdout.flush()
        return {
            "error": error_msg,
            "message": f"Failed to connect: {error_msg}"
        }

# ==================== AGENT 2: EMAIL PARSER ====================
def agent_email_parser(state: AgentState) -> dict:
    """Parse email details"""
    print(f"\n[STEP 4] Extracting email details...")
    sys.stdout.flush()
    
    email_ids = state.get("email_ids", [])
    if not email_ids:
        print(f"✗ No emails to parse")
        sys.stdout.flush()
        return {
            "error": "No emails to parse",
            "message": "No emails found",
            "email_data": []
        }
    
    try:
        gmail_account = state.get("gmail_account")
        gmail_password = state.get("gmail_password")
        
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(gmail_account, gmail_password)
        imap.select('"[Gmail]/Sent Mail"')
        
        email_list = []
        
        for idx, email_id in enumerate(email_ids, 1):
            try:
                status, msg_data = imap.fetch(email_id, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                
                # Extract and clean email fields
                to_email = msg.get('To', 'N/A')
                to_email = extract_email_address(to_email)
                
                subject = msg.get('Subject', 'N/A')
                date = msg.get('Date', 'N/A')
                
                email_list.append({
                    "Email ID": email_id.decode(),
                    "To": to_email,
                    "Subject": subject,
                    "Date": date
                })
                print(f"  Email {idx}/{len(email_ids)} parsed")
                sys.stdout.flush()
                
            except Exception as e:
                print(f"  Warning: Could not parse email {email_id}")
                sys.stdout.flush()
                continue
        
        imap.close()
        print(f"✓ Extracted {len(email_list)} emails")
        sys.stdout.flush()
        
        return {
            "email_data": email_list,
            "message": f"Extracted {len(email_list)} email details"
        }
        
    except Exception as e:
        error_msg = str(e)
        print(f"✗ Error: {error_msg}")
        sys.stdout.flush()
        return {
            "error": error_msg,
            "message": f"Error parsing emails: {error_msg}",
            "email_data": []
        }

# ==================== AGENT 3: EXCEL MAKER ====================
def agent_excel_maker(state: AgentState) -> dict:
    """Create Excel file"""
    print(f"\n[STEP 5] Creating Excel file...")
    sys.stdout.flush()
    
    email_data = state.get("email_data", [])
    if not email_data:
        print(f"✗ No emails to export")
        sys.stdout.flush()
        return {
            "error": "No emails to export",
            "message": "No data to export",
            "output_path": ""
        }
    
    try:
        output_path = os.path.expanduser("~/Desktop/sent_emails.xlsx")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sent Emails"
        
        # Add headers
        headers = ["Email ID", "To", "Subject", "Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add data
        for row_num, ed in enumerate(email_data, 2):
            for col_num, (key, value) in enumerate(ed.items(), 1):
                ws.cell(row=row_num, column=col_num).value = value
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        
        print(f"✓ Excel file saved at: {output_path}")
        sys.stdout.flush()
        print(f"\n[STEP 6] Workflow completed successfully!")
        sys.stdout.flush()
        
        return {
            "output_path": output_path,
            "message": f"Successfully created Excel with {len(email_data)} emails"
        }
        
    except Exception as e:
        error_msg = str(e)
        print(f"✗ Error creating Excel: {error_msg}")
        sys.stdout.flush()
        return {
            "error": error_msg,
            "message": f"Error: {error_msg}",
            "output_path": ""
        }

# ==================== WORKFLOW GRAPH ====================
def create_workflow():
    """Create LangGraph workflow"""
    workflow = StateGraph(AgentState)
    
    # Add nodes
    workflow.add_node("gmail_extractor", agent_gmail_extractor)
    workflow.add_node("email_parser", agent_email_parser)
    workflow.add_node("excel_maker", agent_excel_maker)
    
    # Add edges
    workflow.set_entry_point("gmail_extractor")
    workflow.add_edge("gmail_extractor", "email_parser")
    workflow.add_edge("email_parser", "excel_maker")
    workflow.add_edge("excel_maker", END)
    
    return workflow.compile()

# ==================== RUN WORKFLOW ====================
def run_email_extraction_workflow(gmail_account: str, gmail_password: str) -> dict:
    """Execute the workflow"""
    
    initial_state: AgentState = {
        "gmail_account": gmail_account,
        "gmail_password": gmail_password,
        "email_ids": [],
        "email_data": [],
        "output_path": "",
        "message": "",
        "error": ""
    }
    
    try:
        workflow = create_workflow()
        final_state = workflow.invoke(initial_state)
        
        return {
            "success": final_state.get("output_path") != "",
            "email_count": len(final_state.get("email_data", [])),
            "output_path": final_state.get("output_path", ""),
            "message": final_state.get("message", ""),
            "error": final_state.get("error", "")
        }
    except Exception as e:
        print(f"Workflow error: {e}")
        sys.stdout.flush()
        return {
            "success": False,
            "email_count": 0,
            "output_path": "",
            "message": f"Workflow error: {e}",
            "error": str(e)
        }

# Legacy: Direct execution
if __name__ == "__main__":
    result = run_email_extraction_workflow(
        gmail_account="your_gmail_id@gmail.com",
        gmail_password="Your password"
    )
    print(f"\n{'='*50}")
    print(f"Final Result: {result}")
